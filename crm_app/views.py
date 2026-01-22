from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login,logout, authenticate
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q, Count, Sum, F, Avg
from django.utils import timezone
from django.http import HttpResponse
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
import csv
from io import BytesIO
from openpyxl import Workbook
from django.views.decorators.csrf import csrf_exempt
from datetime import date, timedelta
from .models import (
    Lead, Course, Group, Room, FollowUp, TrialLesson, 
    KPI, User, Reactivation, LeaveRequest, SalesMessage, SalesMessageRead, Offer
)
from .forms import (
    LeadForm, LeadStatusForm, TrialLessonForm, TrialResultForm,
    FollowUpForm, CustomFollowUpForm, ExcelImportForm, UserCreateForm, UserEditForm,
    CourseForm, RoomForm, GroupForm, LeaveRequestForm,
    LeaveRequestApprovalForm, SalesAbsenceForm, SalesMessageForm, OfferForm
)
from .decorators import role_required, admin_required, manager_or_admin_required
from .services import (
    LeadDistributionService, FollowUpService, GroupService,
    KPIService, ReactivationService, OfferService, GoogleSheetsService
)
try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False
    try:
        from openpyxl import load_workbook
    except ImportError:
        load_workbook = None


def login_view(request):
    if request.user.is_authenticated:
        return redirect('dashboard')
    
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user:
            login(request, user)
            return redirect('dashboard')
        else:
            messages.error(request, 'Login yoki parol noto\'g\'ri')
    
    return render(request, 'login.html')

def logout_view(request):
    logout(request)
    return redirect('login')


def landing_page(request):
    """Marketing landing: loyiha haqida ma'lumot va buyurtma CTA"""
    from django.urls import reverse
    from .telegram_bot import send_telegram_notification
    canonical_url = request.build_absolute_uri(reverse('landing'))
    success = False

    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        phone = request.POST.get('phone', '').strip()
        notes = request.POST.get('notes', '').strip()

        if not name or not phone:
            messages.error(request, "Ism va telefon raqam majburiy.")
        else:
            msg = (
                "🛒 Yangi buyurtma (landing)\n"
                f"👤 {name}\n"
                f"📞 {phone}\n"
                f"📝 {notes or '—'}"
            )
            # Admin/Sales manager chat yoki guruhga yuborish
            recipients = list(User.objects.filter(role__in=['admin', 'sales_manager']))
            sent = 0
            for u in recipients:
                # Agar guruh ID bo'lsa, o'sha guruhga, aks holda shaxsiy chatga
                target = u.telegram_group_id or u.telegram_chat_id
                if target:
                    if send_telegram_notification(target, msg):
                        sent += 1
            success = sent > 0
            if success:
                messages.success(request, "So'rov yuborildi. Tez orada bog'lanamiz.")
            else:
                messages.warning(request, "So'rov qabul qilindi, lekin Telegram yuborilmadi (kontaktlar sozlanmagan).")

    context = {
        'canonical_url': canonical_url,
        'success': success,
    }
    return render(request, 'landing.html', context)

@login_required
def dashboard(request):
    try:
        context = {}
        now = timezone.now()
        
        if request.user.is_admin or request.user.is_sales_manager:
            # Admin/Manager dashboard
            overdue_followups_queryset = FollowUpService.get_overdue_followups_prioritized()
            
            context.update({
            'total_leads': Lead.objects.count(),
            'new_leads_today': Lead.objects.filter(
                created_at__date=timezone.now().date(),
                status='new'
            ).count(),
            'total_sales': User.objects.filter(role='sales', is_active_sales=True).count(),
            'active_groups': Group.objects.filter(is_active=True).count(),
            'overdue_followups': overdue_followups_queryset.count(),
            'overdue_followups_list': overdue_followups_queryset.select_related(
                'lead', 'sales', 'lead__assigned_sales', 'lead__interested_course'
            )[:10],  # Eng qadimgi 10 tasi
                'overdue_stats': FollowUpService.get_overdue_statistics(),
            })
        else:
            # Sales dashboard
            sales = request.user
            overdue_followups_queryset = FollowUpService.get_overdue_followups_prioritized(sales)
            
            # Bugungi KPI
            today = timezone.now().date()
            today_kpi = KPIService.calculate_daily_kpi(sales, today)
            
            # Oxirgi 7 kunlik KPI
            from datetime import timedelta
            last_7_days_kpi = []
            for i in range(7):
                date = today - timedelta(days=i)
                kpi = KPI.objects.filter(sales=sales, date=date).first()
                if not kpi:
                    kpi = KPIService.calculate_daily_kpi(sales, date)
                last_7_days_kpi.append({
                    'date': date,
                    'kpi': kpi
                })
            
            # Reyting (conversion_rate bo'yicha)
            ranking = KPIService.get_sales_ranking(sales, period='month', metric='conversion_rate')
            
            # Trend taqqoslash
            trend_contacts = KPIService.get_trend_comparison(sales, days=7, metric='daily_contacts')
            trend_conversion = KPIService.get_trend_comparison(sales, days=7, metric='conversion_rate')
            
            context.update({
                'my_leads': Lead.objects.filter(assigned_sales=sales).count(),
                'today_followups': FollowUpService.get_today_followups(sales).count(),
                'overdue_followups': overdue_followups_queryset.count(),
                'overdue_followups_list': overdue_followups_queryset.select_related(
                    'lead', 'lead__interested_course'
                )[:10],  # Eng qadimgi 10 tasi
                'is_blocked': FollowUpService.check_sales_blocked(sales),
                'today_kpi': today_kpi,
                'last_7_days_kpi': last_7_days_kpi,
                'ranking': ranking,
                'trend_contacts': trend_contacts,
                'trend_conversion': trend_conversion,
            })
        
        return render(request, 'dashboard.html', context)
    except Exception as e:
        messages.error(request, f'Dashboard yuklashda xatolik: {str(e)}')
        import traceback
        traceback.print_exc()
        return redirect('leads_list')


# ============ LEAD MANAGEMENT ============

@login_required
@role_required('admin', 'sales_manager', 'sales')
def leads_list(request):
    try:
        queryset = Lead.objects.select_related('assigned_sales', 'interested_course').all()
        
        if request.user.is_sales:
            queryset = queryset.filter(assigned_sales=request.user)
        
        # Filtrlash
        source_filter = request.GET.get('source')
        if source_filter:
            queryset = queryset.filter(source=source_filter)
        
        search = request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(name__icontains=search) | Q(phone__icontains=search)
            )
        
        # Statuslar bo'yicha guruhlash (Kanban board uchun)
        now = timezone.now()
        leads_by_status = {}
        for status_code, status_name in Lead.STATUS_CHOICES:
            status_leads = queryset.filter(status=status_code).prefetch_related(
                'followups',
                'trials__group__course',
                'trials__group',
                'trials__room'
            ).order_by('-created_at')
            # Har bir lid uchun overdue follow-up borligini tekshirish
            for lead in status_leads:
                lead.has_overdue_followup = lead.followups.filter(
                    completed=False, 
                    due_date__lt=now
                ).exists()
                # Eng yaqin sinov darsini topish (natija bo'lmagan yoki kelajakdagi)
                upcoming_trial = lead.trials.filter(
                    result__in=['', 'attended', 'not_attended']
                ).order_by('date', 'time').first()
                if not upcoming_trial:
                    # Agar kelajakdagi bo'lmasa, barcha sinov darslardan eng yaqinini olish
                    upcoming_trial = lead.trials.all().order_by('date', 'time').first()
                lead.upcoming_trial = upcoming_trial
            leads_by_status[status_code] = {
                'name': status_name,
                'leads': status_leads,
                'count': status_leads.count()
            }
        
        context = {
            'leads_by_status': leads_by_status,
            'statuses': Lead.STATUS_CHOICES,
            'sources': Lead.SOURCE_CHOICES,
            'total_leads': queryset.count(),
        }
        return render(request, 'leads/list.html', context)
    except Exception as e:
        messages.error(request, f'Lidlar ro\'yxatini yuklashda xatolik: {str(e)}')
        import traceback
        traceback.print_exc()
        return redirect('dashboard')


@login_required
@role_required('admin', 'sales_manager', 'sales')
def leads_table(request):
    """Lidlar jadvali (filtr + CSV eksport)"""
    try:
        leads = Lead.objects.select_related('assigned_sales', 'interested_course').all()

        status = request.GET.get('status') or ''
        source = request.GET.get('source') or ''
        sales_id = request.GET.get('sales') or ''
        course_id = request.GET.get('course') or ''
        date_from = request.GET.get('date_from') or ''
        date_to = request.GET.get('date_to') or ''

        # Sales o'z lidlarini ko'radi
        if request.user.is_sales:
            leads = leads.filter(assigned_sales=request.user)

        if status:
            leads = leads.filter(status=status)
        if source:
            leads = leads.filter(source=source)
        if sales_id:
            leads = leads.filter(assigned_sales_id=sales_id)
        if course_id:
            leads = leads.filter(interested_course_id=course_id)
        if date_from:
            leads = leads.filter(created_at__date__gte=date_from)
        if date_to:
            leads = leads.filter(created_at__date__lte=date_to)

        # Export Excel (pagination dan oldin)
        if request.GET.get('export') == 'excel':
            wb = Workbook()
            ws = wb.active
            ws.title = "Leads"
            headers = ['Ism', 'Telefon', "Qo'shimcha telefon", 'Status', 'Manba', 'Kurs', 'Sotuvchi', 'Yaratilgan']
            ws.append(headers)
            for lead in leads:
                ws.append([
                    lead.name,
                    lead.phone,
                    getattr(lead, 'secondary_phone', '') or '',
                    lead.get_status_display(),
                    lead.get_source_display(),
                    lead.interested_course.name if lead.interested_course else '',
                    lead.assigned_sales.username if lead.assigned_sales else '',
                    timezone.localtime(lead.created_at).strftime('%d.%m.%Y %H:%M'),
                ])

            bio = BytesIO()
            wb.save(bio)
            bio.seek(0)
            filename = f"leads_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            response = HttpResponse(
                bio.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response

        # Pagination qo'shish
        paginator = Paginator(leads.order_by('-created_at'), 50)  # 50 ta per page
        page = request.GET.get('page', 1)
        
        try:
            leads_page = paginator.page(page)
        except PageNotAnInteger:
            leads_page = paginator.page(1)
        except EmptyPage:
            leads_page = paginator.page(paginator.num_pages)

        context = {
            'leads': leads_page,  # Paginated queryset
            'paginator': paginator,
            'page_obj': leads_page,
            'status_filter': status,
            'source_filter': source,
            'sales_filter': sales_id,
            'course_filter': course_id,
            'date_from': date_from,
            'date_to': date_to,
            'sales_list': User.objects.filter(role='sales', is_active_sales=True),
            'course_list': Course.objects.filter(is_active=True),
            'statuses': Lead.STATUS_CHOICES,
            'sources': Lead.SOURCE_CHOICES,
        }
        return render(request, 'leads/table.html', context)
    except Exception as e:
        messages.error(request, f'Lidlar jadvalini yuklashda xatolik: {str(e)}')
        import traceback
        traceback.print_exc()
        return redirect('leads_list')


@login_required
@role_required('admin', 'sales_manager', 'sales')
def lead_detail(request, pk):
    try:
        lead = get_object_or_404(Lead, pk=pk)
        
        # Faqat o'z lidlari bilan ishlash
        if request.user.is_sales and lead.assigned_sales != request.user:
            messages.error(request, "Sizda bu lidni ko'rish huquqi yo'q")
            return redirect('leads_list')
        
        if request.method == 'POST':
            # Status o'zgartirish
            if 'status_form' in request.POST:
                form = LeadStatusForm(request.POST, instance=lead)
                if form.is_valid():
                    form.save()
                    messages.success(request, 'Status yangilandi')
                    return redirect('lead_detail', pk=pk)
            # Lid ma'lumotlarini o'zgartirish
            elif 'edit_form' in request.POST:
                edit_form = LeadForm(request.POST, instance=lead)
                if edit_form.is_valid():
                    edit_form.save()
                    messages.success(request, 'Lid ma\'lumotlari yangilandi')
                    return redirect('lead_detail', pk=pk)
            # Custom follow-up yaratish
            elif 'create_custom_followup' in request.POST:
                custom_followup_form = CustomFollowUpForm(request.POST)
                if custom_followup_form.is_valid():
                    due_datetime = custom_followup_form.cleaned_data['due_date']
                    notes = custom_followup_form.cleaned_data.get('notes', '')
                    
                    # Timezone-aware qilish (agar naive bo'lsa)
                    if due_datetime and timezone.is_naive(due_datetime):
                        due_datetime = timezone.make_aware(due_datetime)
                    
                    # Ish vaqtini tekshirish va moslashtirish
                    sales = request.user
                    now = timezone.now()
                    
                    # Agar ish vaqtlari belgilanmagan bo'lsa, oddiy yaratish
                    if not sales.work_start_time or not sales.work_end_time:
                        try:
                            followup = FollowUp.objects.create(
                                lead=lead,
                                sales=sales,
                                due_date=due_datetime,
                                notes=f"Sotuvchi tomonidan belgilangan: {notes}" if notes else "Sotuvchi tomonidan belgilangan"
                            )
                            from .tasks import send_followup_created_notification
                            send_followup_created_notification.delay(followup.id)
                            messages.success(request, f'Vazifa yaratildi: {due_datetime.strftime("%d.%m.%Y %H:%M")}')
                            return redirect('lead_detail', pk=pk)
                        except Exception as e:
                            messages.error(request, f'Vazifa yaratishda xatolik: {str(e)}')
                            import traceback
                            traceback.print_exc()
                            custom_followup_form = CustomFollowUpForm(request.POST)
                    else:
                        # Ish vaqtini tekshirish
                        due_time = due_datetime.time()
                        if due_time < sales.work_start_time or due_time > sales.work_end_time:
                            messages.error(request, f'Belgilangan vaqt ({due_time.strftime("%H:%M")}) ish vaqti tashqarisida. Ish vaqti: {sales.work_start_time.strftime("%H:%M")} - {sales.work_end_time.strftime("%H:%M")}')
                            custom_followup_form = CustomFollowUpForm(request.POST)
                        else:
                            # Ish vaqtini moslashtirish (ish kunlarini tekshirish)
                            # Timezone-aware qilish (agar naive bo'lsa)
                            if timezone.is_naive(due_datetime):
                                due_datetime = timezone.make_aware(due_datetime)
                            
                            # Delay hisoblash (timezone-aware datetime'lar orasida)
                            delay = due_datetime - now
                            
                            try:
                                adjusted_due_date = FollowUpService.calculate_work_hours_due_date(
                                    sales,
                                    now,
                                    delay
                                )
                                
                                try:
                                    followup = FollowUp.objects.create(
                                        lead=lead,
                                        sales=sales,
                                        due_date=adjusted_due_date,
                                        notes=f"Sotuvchi tomonidan belgilangan: {notes}" if notes else "Sotuvchi tomonidan belgilangan"
                                    )
                                    
                                    # Notification yuborish
                                    from .tasks import send_followup_created_notification
                                    send_followup_created_notification.delay(followup.id)
                                    
                                    messages.success(request, f'Vazifa yaratildi: {adjusted_due_date.strftime("%d.%m.%Y %H:%M")}')
                                    return redirect('lead_detail', pk=pk)
                                except Exception as e:
                                    messages.error(request, f'Vazifa yaratishda xatolik: {str(e)}')
                                    import traceback
                                    traceback.print_exc()
                                    custom_followup_form = CustomFollowUpForm(request.POST)
                            except Exception as e:
                                messages.error(request, f'Ish vaqtini hisoblashda xatolik: {str(e)}')
                                import traceback
                                traceback.print_exc()
                                custom_followup_form = CustomFollowUpForm(request.POST)
                else:
                    # Form validation xatolari
                    for field, errors in custom_followup_form.errors.items():
                        for error in errors:
                            messages.error(request, f'{custom_followup_form.fields[field].label}: {error}')
        else:
            form = LeadStatusForm(instance=lead)
            edit_form = LeadForm(instance=lead)
            custom_followup_form = CustomFollowUpForm()
        
        # Faol takliflarni olish
        from .services import OfferService
        active_offers = OfferService.active_for_lead(lead, channel='all')
        
        # Kurs sotuv scriptini olish
        sales_script = None
        sales_script_course_name = None
        if lead.interested_course:
            sales_script = lead.interested_course.sales_script
            sales_script_course_name = lead.interested_course.name
        else:
            # Umumiy sotuv scriptini qidirish
            general_course = Course.objects.filter(name__icontains='umumiy').first()
            if general_course:
                sales_script = general_course.sales_script
                sales_script_course_name = "Umumiy"
        
        # Custom follow-up form (agar POST bo'lmagan bo'lsa yoki xatolik bo'lsa)
        if 'custom_followup_form' not in locals():
            custom_followup_form = CustomFollowUpForm()
        
        context = {
            'lead': lead,
            'form': form,
            'edit_form': edit_form,
            'custom_followup_form': custom_followup_form,
            'followups': lead.followups.select_related('sales').all().order_by('-due_date'),
            'trials': lead.trials.select_related('group', 'group__course', 'room').all().order_by('-date'),
            'active_offers': active_offers,
            'sales_script': sales_script,
            'sales_script_course_name': sales_script_course_name,
        }
        return render(request, 'leads/detail.html', context)
    except Exception as e:
        messages.error(request, f'Lid ma\'lumotlarini yuklashda xatolik: {str(e)}')
        import traceback
        traceback.print_exc()
        return redirect('leads_list')


@login_required
@role_required('admin', 'sales_manager', 'sales')
def lead_create(request):
    if request.method == 'POST':
        form = LeadForm(request.POST, user=request.user)
        if form.is_valid():
            lead = form.save(commit=False)
            
            # Agar sotuvchi o'zi lid qo'shgan bo'lsa
            if request.user.is_sales:
                # Agar assigned_sales tanlanmagan bo'lsa, o'ziga biriktirish
                if not lead.assigned_sales:
                    lead.assigned_sales = request.user
                # Agar tanlangan bo'lsa, tanlangan sotuvchiga biriktirish
                lead.save()
                from .tasks import send_new_lead_notification
                send_new_lead_notification.delay(lead.id)
            elif not lead.assigned_sales:
                # Admin/Manager uchun avtomatik taqsimlash
                LeadDistributionService.distribute_leads([lead])
            else:
                # Agar qo'lda biriktirilgan bo'lsa, saqlash va notification yuborish
                lead.save()
                from .tasks import send_new_lead_notification
                send_new_lead_notification.delay(lead.id)
            
            messages.success(request, 'Lid qo\'shildi')
            return redirect('lead_detail', pk=lead.pk)
    else:
        form = LeadForm(user=request.user)
    
    return render(request, 'leads/create.html', {'form': form})


@login_required
@role_required('admin', 'sales_manager')
def lead_assign(request, pk):
    lead = get_object_or_404(Lead, pk=pk)
    
    if request.method == 'POST':
        sales_id = request.POST.get('sales_id')
        new_sales = get_object_or_404(User, pk=sales_id, role='sales')
        old_sales = lead.assigned_sales  # Eski sotuvchini saqlash
        
        # Lidni yangi sotuvchiga biriktirish
        lead.assigned_sales = new_sales
        lead.save()
        
        # Eski sotuvchining bajarilmagan follow-up'larini yangi sotuvchiga o'tkazish
        pending_count = 0
        if old_sales:
            pending_followups = FollowUp.objects.filter(
                lead=lead,
                sales=old_sales,
                completed=False
            )
            pending_count = pending_followups.count()
            
            # Follow-up'larni yangi sotuvchiga o'tkazish
            if pending_count > 0:
                pending_followups.update(sales=new_sales)
            
            # Eski sotuvchiga xabar yuborish (agar telegram_chat_id bo'lsa)
            if old_sales.telegram_chat_id:
                from .telegram_bot import send_telegram_notification
                message = (
                    f"🔔 Lid o'tkazildi\n\n"
                    f"👤 Lid: {lead.name}\n"
                    f"📞 Telefon: {lead.phone}\n"
                    f"➡️ Yangi sotuvchi: {new_sales.username}\n"
                )
                if pending_count > 0:
                    message += f"📋 {pending_count} ta bajarilmagan follow-up yangi sotuvchiga o'tkazildi."
                else:
                    message += f"✅ Barcha follow-up'lar bajarilgan edi."
                
                send_telegram_notification(old_sales.telegram_chat_id, message)
        
        # Yangi sotuvchiga notification yuborish
        from .tasks import send_new_lead_notification
        send_new_lead_notification.delay(lead.id)
        
        messages.success(request, f'Lid {new_sales.username} ga biriktirildi')
        if pending_count > 0:
            messages.info(request, f'{pending_count} ta bajarilmagan follow-up yangi sotuvchiga o\'tkazildi')
        return redirect('lead_detail', pk=pk)
    
    sales_list = User.objects.filter(role='sales', is_active_sales=True)
    return render(request, 'leads/assign.html', {
        'lead': lead,
        'sales_list': sales_list
    })


@login_required
@role_required('admin', 'sales_manager', 'sales')
def excel_import(request):
    if request.method == 'POST':
        form = ExcelImportForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                leads = []
                
                if PANDAS_AVAILABLE:
                    # Pandas bilan
                    df = pd.read_excel(request.FILES['file'])
                    for _, row in df.iterrows():
                        # Duplicate tekshirish
                        phone = str(row.get('phone', '')).strip()
                        if not phone or Lead.objects.filter(phone=phone).exists():
                            continue
                        
                        # Source ustunini o'qish
                        source_value = str(row.get('source', '')).strip().lower()
                        # Source'ni Lead modelidagi SOURCE_CHOICES ga moslashtirish
                        source = 'excel'  # Default
                        if source_value:
                            # Mapping
                            source_mapping = {
                                'instagram': 'instagram',
                                'telegram': 'telegram',
                                'youtube': 'youtube',
                                'organic': 'organic',
                                'form': 'form',
                                'excel': 'excel',
                                'google_sheets': 'google_sheets',
                            }
                            source = source_mapping.get(source_value, 'excel')
                        
                        # Interested course ni qidirish va biriktirish (case-insensitive)
                        interested_course = None
                        course_name = None
                        
                        # Mumkin bo'lgan ustun nomlari
                        course_key_variants = ['course', 'kurs', 'interested_course', 'qiziqqan kurs']
                        row_keys_lower = {str(k).lower().strip(): k for k in row.keys()}
                        
                        for variant in course_key_variants:
                            if variant in row_keys_lower:
                                found_key = row_keys_lower[variant]
                                course_name = str(row.get(found_key, '')).strip()
                                if course_name:
                                    break
                        
                        # Agar variantlardan topilmasa, default ustun nomidan qidirish
                        if not course_name:
                            course_name = str(row.get('course', row.get('interested_course', ''))).strip()
                        
                        if course_name:
                            try:
                                interested_course = Course.objects.filter(
                                    name__icontains=course_name
                                ).first()
                            except:
                                pass
                        
                        # Secondary phone
                        secondary_phone = str(row.get('secondary_phone', '')).strip()
                        if secondary_phone:
                            secondary_phone = secondary_phone.replace('+', '').replace(' ', '').replace('-', '').replace('(', '').replace(')', '')
                        
                        lead = Lead(
                            name=str(row.get('name', '')).strip(),
                            phone=phone,
                            secondary_phone=secondary_phone if secondary_phone else None,
                            source=source,
                            interested_course=interested_course,
                        )
                        leads.append(lead)
                elif load_workbook:
                    # Openpyxl bilan (pandassiz)
                    wb = load_workbook(request.FILES['file'])
                    ws = wb.active
                    
                    # Birinchi qatorni header sifatida o'qish
                    headers = [cell.value for cell in ws[1]]
                    name_col = None
                    phone_col = None
                    source_col = None
                    course_col = None
                    secondary_phone_col = None
                    
                    for idx, header in enumerate(headers, 1):
                        header_str = str(header or '').lower()
                        if 'name' in header_str:
                            name_col = idx
                        if 'phone' in header_str:
                            phone_col = idx
                        if 'source' in header_str:
                            source_col = idx
                        if 'course' in header_str or 'kurs' in header_str or 'interested_course' in header_str:
                            course_col = idx
                        if 'secondary_phone' in header_str or 'qo\'shimcha' in header_str:
                            secondary_phone_col = idx
                    
                    if not name_col or not phone_col:
                        messages.error(request, 'Excel faylda "name" va "phone" ustunlari topilmadi')
                        return render(request, 'leads/import.html', {'form': form})
                    
                    # Ma'lumotlarni o'qish
                    for row in ws.iter_rows(min_row=2, values_only=False):
                        phone = str(row[phone_col - 1].value or '').strip()
                        name = str(row[name_col - 1].value or '').strip()
                        
                        if not phone or not name:
                            continue
                        
                        # Duplicate tekshirish
                        if Lead.objects.filter(phone=phone).exists():
                            continue
                        
                        # Source ustunini o'qish
                        source = 'excel'  # Default
                        if source_col:
                            source_value = str(row[source_col - 1].value or '').strip().lower()
                            if source_value:
                                # Mapping
                                source_mapping = {
                                    'instagram': 'instagram',
                                    'telegram': 'telegram',
                                    'youtube': 'youtube',
                                    'organic': 'organic',
                                    'form': 'form',
                                    'excel': 'excel',
                                    'google_sheets': 'google_sheets',
                                }
                                source = source_mapping.get(source_value, 'excel')
                        
                        # Interested course ni qidirish va biriktirish
                        interested_course = None
                        if course_col:
                            course_name = str(row[course_col - 1].value or '').strip()
                            if course_name:
                                try:
                                    interested_course = Course.objects.filter(
                                        name__icontains=course_name
                                    ).first()
                                except:
                                    pass
                        
                        # Secondary phone
                        secondary_phone = None
                        if secondary_phone_col:
                            secondary_phone = str(row[secondary_phone_col - 1].value or '').strip()
                            if secondary_phone:
                                secondary_phone = secondary_phone.replace('+', '').replace(' ', '').replace('-', '').replace('(', '').replace(')', '')
                        
                        lead = Lead(
                            name=name,
                            phone=phone,
                            secondary_phone=secondary_phone if secondary_phone else None,
                            source=source,
                            interested_course=interested_course,
                        )
                        leads.append(lead)
                else:
                    messages.error(request, 'Excel fayllarni o\'qish uchun pandas yoki openpyxl kerak')
                    return render(request, 'leads/import.html', {'form': form})
                
                # Taqsimlash (distribute_leads ichida save va notification yuboriladi)
                if leads:
                    # MUHIM: Lidlarni avval save qilmaslik! 
                    # distribute_leads ichida assigned_sales bilan birga save qiladi
                    # Bu signal'ni to'g'ri ishlashi va follow-up yaratilishi uchun zarur
                    LeadDistributionService.distribute_leads(leads)
                    
                    messages.success(request, f'{len(leads)} ta lid import qilindi va taqsimlandi')
                else:
                    messages.warning(request, 'Yangi lid topilmadi')
                
                return redirect('leads_list')
            except Exception as e:
                messages.error(request, f'Xatolik: {str(e)}')
    else:
        form = ExcelImportForm()
    
    return render(request, 'leads/import.html', {'form': form})


@login_required
@role_required('admin', 'sales_manager', 'sales')
def google_sheets_manual_import(request):
    """Google Sheets'dan qo'lda yangi lidlarni import qilish (tugma bosilganda)"""
    from django.conf import settings
    
    try:
        # .env dan konfiguratsiyani olish
        spreadsheet_id = getattr(settings, 'GOOGLE_SHEETS_SPREADSHEET_ID', None)
        worksheets = getattr(
            settings,
            'GOOGLE_SHEETS_WORKSHEETS',
            [getattr(settings, 'GOOGLE_SHEETS_WORKSHEET_NAME', 'Sheet1')]
        )
        
        if not spreadsheet_id:
            messages.error(request, 'Google Sheets ID sozlanmagan! .env faylda GOOGLE_SHEETS_SPREADSHEET_ID ni sozlang.')
            return redirect('leads_list')
        
        summary = {'imported': 0, 'skipped': 0, 'errors': []}
        
        # Bir nechta worksheetlar bo'yicha import qilish
        for worksheet_name in worksheets:
            result = GoogleSheetsService.import_new_leads(
                spreadsheet_id=spreadsheet_id,
                worksheet_name=worksheet_name
            )
            summary['imported'] += result.get('imported', 0)
            summary['skipped'] += result.get('skipped', 0)
            summary['errors'].extend([f"{worksheet_name}: {err}" for err in result.get('errors', [])])
        
        # Natijalarni ko'rsatish
        if summary['imported'] > 0:
            messages.success(request, f"✅ {summary['imported']} ta yangi lid muvaffaqiyatli import qilindi!")
        elif summary['skipped'] > 0:
            messages.info(request, f"ℹ️ Yangi lid topilmadi. {summary['skipped']} ta lid allaqachon mavjud.")
        else:
            messages.info(request, "ℹ️ Yangi lid topilmadi.")
        
        if summary['errors']:
            for error in summary['errors'][:3]:  # Faqat birinchi 3 ta xatoni ko'rsatish
                messages.warning(request, f"⚠️ {error}")
            if len(summary['errors']) > 3:
                messages.warning(request, f"... va yana {len(summary['errors']) - 3} ta xato")
        
    except Exception as e:
        messages.error(request, f'❌ Xatolik yuz berdi: {str(e)}')
    
    return redirect('leads_list')


# ============ FOLLOW-UP MANAGEMENT ============

@login_required
@role_required('admin', 'sales_manager', 'sales')
def followups_today(request):
    # Bajarilmagan follow-up'larni olish
    followups = FollowUpService.get_today_followups(
        sales=request.user if request.user.is_sales else None
    )
    active_offers = OfferService.active_offers(channel='followup')
    
    # Statistikalar - to'g'ri hisoblash
    # get_today_followups allaqachon completed=False filter qilgan
    # Shuning uchun stats'da bajarilganlarni alohida qidirish kerak
    now = timezone.now()
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    today_end = now.replace(hour=23, minute=59, second=59, microsecond=999999)
    
    # Barcha bugungi follow-up'lar (bajarilganlari ham) - stats uchun
    all_today_followups = FollowUp.objects.select_related(
        'lead', 'sales', 'lead__assigned_sales', 'lead__interested_course'
    ).filter(
        due_date__gte=today_start,
        due_date__lte=today_end
    )
    
    if request.user.is_sales:
        all_today_followups = all_today_followups.filter(sales=request.user)
    
    # Statistikalar
    stats = {
        'total': all_today_followups.count(),  # Barcha bugungi follow-up'lar
        'completed': all_today_followups.filter(completed=True).count(),  # Bajarilganlar
        'pending': followups.filter(is_overdue=False).count(),  # Bajarilmagan va overdue emas
        'overdue': followups.filter(is_overdue=True).count(),  # Bajarilmagan va overdue
    }
    stats['completion_rate'] = (stats['completed'] / stats['total'] * 100) if stats['total'] > 0 else 0
    
    if request.method == 'POST':
        followup_id = request.POST.get('followup_id')
        followup = get_object_or_404(FollowUp, pk=followup_id)
        if request.user.is_sales and followup.sales != request.user:
            messages.error(request, "Sizda bu follow-upni bajarish huquqi yo'q")
        else:
            # Ish vaqti va ruxsat tekshirish (faqat sales uchun)
            # Eslatma: request.user ish vaqtida bo'lishi kerak, followup.sales emas
            if request.user.is_sales:
                # Aniqroq xabar uchun alohida tekshirish
                if not request.user.is_active_sales:
                    messages.error(request, "Siz faol sotuvchi emassiz.")
                    return redirect('followups_today')
                
                # Ruxsat tekshirish
                from django.apps import apps
                LeaveRequest = apps.get_model('crm_app', 'LeaveRequest')
                now = timezone.now()
                active_leave = LeaveRequest.objects.filter(
                    sales=request.user,
                    status='approved',
                    start_date__lte=now.date(),
                    end_date__gte=now.date()
                ).first()
                
                is_on_leave = False
                if active_leave:
                    if active_leave.start_time and active_leave.end_time:
                        current_time = now.time()
                        if active_leave.start_time <= current_time <= active_leave.end_time:
                            is_on_leave = True
                    else:
                        is_on_leave = True
                
                if is_on_leave or request.user.is_on_leave:
                    messages.error(request, "Siz hozir ruxsat olgansiz. Follow-up'ni faqat ish vaqtida bajarilgan deb belgilash mumkin.")
                    return redirect('followups_today')
                
                # Ish vaqti tekshirish
                if not request.user.is_working_at_time(now):
                    messages.error(request, "Siz hozir ish vaqtida emassiz. Follow-up'ni faqat ish vaqtida bajarilgan deb belgilash mumkin.")
                    return redirect('followups_today')
            
        try:
            # check_work_hours=False, chunki biz allaqachon yuqorida tekshirdik
            followup.mark_completed(check_work_hours=False)
            messages.success(request, 'Follow-up bajarildi')
        except Exception as e:
            from django.core.exceptions import ValidationError
            if isinstance(e, ValidationError):
                messages.error(request, str(e))
            else:
                messages.error(request, f'Xatolik: {str(e)}')
        return redirect('followups_today')
    
    # Pagination qo'shish
    paginator = Paginator(followups, 30)  # 30 ta per page
    page = request.GET.get('page', 1)
    
    try:
        followups_page = paginator.page(page)
    except PageNotAnInteger:
        followups_page = paginator.page(1)
    except EmptyPage:
        followups_page = paginator.page(paginator.num_pages)
    
    return render(request, 'followups/today.html', {
        'followups': followups_page,  # Paginated
        'paginator': paginator,
        'page_obj': followups_page,
        'active_offers': active_offers,
        'stats': stats,
    })


@login_required
@role_required('admin', 'sales_manager', 'sales')
def overdue_followups_list(request):
    """Overdue follow-up'lar ro'yxati"""
    sales = request.user if request.user.is_sales else None
    active_offers = OfferService.active_offers(channel='followup')
    
    # Filterlar
    sales_filter = request.GET.get('sales')
    age_filter = request.GET.get('age')  # <1h, 1-6h, 6-24h, >24h
    
    overdue = FollowUpService.get_overdue_followups_prioritized(sales).filter(completed=False)
    
    if sales_filter and (request.user.is_admin or request.user.is_sales_manager):
        overdue = overdue.filter(sales_id=sales_filter)
    
    if age_filter:
        now = timezone.now()
        if age_filter == '<1h':
            overdue = overdue.filter(due_date__gte=now - timedelta(hours=1))
        elif age_filter == '1-6h':
            overdue = overdue.filter(
                due_date__lt=now - timedelta(hours=1),
                due_date__gte=now - timedelta(hours=6)
            )
        elif age_filter == '6-24h':
            overdue = overdue.filter(
                due_date__lt=now - timedelta(hours=6),
                due_date__gte=now - timedelta(days=1)
            )
        elif age_filter == '>24h':
            overdue = overdue.filter(due_date__lt=now - timedelta(days=1))
    
    # Statistics
    stats = FollowUpService.get_overdue_statistics(sales)
    
    # Sales list (for filter)
    sales_list = User.objects.filter(role='sales', is_active_sales=True) if (request.user.is_admin or request.user.is_sales_manager) else None
    
    # Pagination qo'shish
    overdue_queryset = overdue.select_related('lead', 'sales', 'lead__interested_course')
    paginator = Paginator(overdue_queryset, 30)  # 30 ta per page
    page = request.GET.get('page', 1)
    
    try:
        overdue_page = paginator.page(page)
    except PageNotAnInteger:
        overdue_page = paginator.page(1)
    except EmptyPage:
        overdue_page = paginator.page(paginator.num_pages)
    
    context = {
        'overdue_followups': overdue_page,  # Paginated
        'paginator': paginator,
        'page_obj': overdue_page,
        'stats': stats,
        'sales_list': sales_list,
        'current_sales_filter': sales_filter,
        'current_age_filter': age_filter,
        'active_offers': active_offers,
    }
    
    return render(request, 'followups/overdue.html', context)


@login_required
@role_required('admin', 'sales_manager', 'sales')
def overdue_followup_complete(request, followup_id):
    """Overdue follow-up'ni bajarilgan deb belgilash"""
    followup = get_object_or_404(FollowUp, pk=followup_id)
    
    if request.user.is_sales and followup.sales != request.user:
        messages.error(request, "Sizda bu follow-upni bajarish huquqi yo'q")
    else:
        # Ish vaqti va ruxsat tekshirish (faqat sales uchun)
        # Eslatma: request.user ish vaqtida bo'lishi kerak, followup.sales emas
        if request.user.is_sales:
            # Aniqroq xabar uchun alohida tekshirish
            if not request.user.is_active_sales:
                messages.error(request, "Siz faol sotuvchi emassiz.")
                return redirect('overdue_followups_list')
            
            # Ruxsat tekshirish
            from django.apps import apps
            LeaveRequest = apps.get_model('crm_app', 'LeaveRequest')
            now = timezone.now()
            active_leave = LeaveRequest.objects.filter(
                sales=request.user,
                status='approved',
                start_date__lte=now.date(),
                end_date__gte=now.date()
            ).first()
            
            is_on_leave = False
            if active_leave:
                if active_leave.start_time and active_leave.end_time:
                    current_time = now.time()
                    if active_leave.start_time <= current_time <= active_leave.end_time:
                        is_on_leave = True
                else:
                    is_on_leave = True
            
            if is_on_leave or request.user.is_on_leave:
                messages.error(request, "Siz hozir ruxsat olgansiz. Follow-up'ni faqat ish vaqtida bajarilgan deb belgilash mumkin.")
                return redirect('overdue_followups_list')
            
            # Ish vaqti tekshirish
            if not request.user.is_working_at_time(now):
                messages.error(request, "Siz hozir ish vaqtida emassiz. Follow-up'ni faqat ish vaqtida bajarilgan deb belgilash mumkin.")
                return redirect('overdue_followups_list')
        
        try:
            # check_work_hours=False, chunki biz allaqachon yuqorida tekshirdik
            followup.mark_completed(check_work_hours=False)
            messages.success(request, 'Follow-up bajarildi')
        except Exception as e:
            from django.core.exceptions import ValidationError
            if isinstance(e, ValidationError):
                messages.error(request, str(e))
            else:
                messages.error(request, f'Xatolik: {str(e)}')
    
    return redirect('overdue_followups_list')


@login_required
@role_required('admin', 'sales_manager')
def bulk_reschedule_overdue(request):
    """Bir nechta overdue'larni qayta rejalashtirish"""
    if request.method == 'POST':
        followup_ids = request.POST.getlist('followup_ids')
        hours_ahead = int(request.POST.get('hours_ahead', 2))
        
        if not followup_ids:
            messages.error(request, "Hech qanday follow-up tanlanmagan")
        else:
            count = 0
            for followup_id in followup_ids:
                followup = get_object_or_404(FollowUp, pk=followup_id)
                if FollowUpService.auto_reschedule_overdue(followup, hours_ahead):
                    count += 1
            
            messages.success(request, f'{count} ta follow-up qayta rejalashtirildi')
    
    return redirect('overdue_followups_list')


@login_required
@role_required('admin', 'sales_manager')
def bulk_reassign_overdue(request):
    """Bir nechta overdue'larni boshqa sotuvchiga o'tkazish"""
    if request.method == 'POST':
        followup_ids = request.POST.getlist('followup_ids')
        new_sales_id = request.POST.get('new_sales_id')
        
        if not followup_ids:
            messages.error(request, "Hech qanday follow-up tanlanmagan")
        elif not new_sales_id:
            messages.error(request, "Sotuvchi tanlanmagan")
        else:
            new_sales = get_object_or_404(User, pk=new_sales_id, role='sales')
            count = 0
            for followup_id in followup_ids:
                followup = get_object_or_404(FollowUp, pk=followup_id)
                if FollowUpService.reassign_overdue_followup(followup, new_sales):
                    count += 1
            
            messages.success(request, f'{count} ta follow-up {new_sales.username} ga o\'tkazildi')
    
    return redirect('overdue_followups_list')


@login_required
@role_required('admin', 'sales_manager')
def bulk_complete_overdue(request):
    """Bir nechta overdue'larni bajarilgan deb belgilash"""
    if request.method == 'POST':
        followup_ids = request.POST.getlist('followup_ids')
        
        if not followup_ids:
            messages.error(request, "Hech qanday follow-up tanlanmagan")
        else:
            count = 0
            skipped = 0
            for followup_id in followup_ids:
                followup = get_object_or_404(FollowUp, pk=followup_id)
                # Admin/Manager uchun ish vaqti tekshirilmaydi
                try:
                    followup.mark_completed(check_work_hours=False)
                    count += 1
                except Exception as e:
                    skipped += 1
            
            if count > 0:
                messages.success(request, f'{count} ta follow-up bajarilgan deb belgilandi')
            if skipped > 0:
                messages.warning(request, f'{skipped} ta follow-up o\'tkazib yuborildi')
    
    return redirect('overdue_followups_list')


@login_required
@admin_required
def bulk_delete_overdue(request):
    """Bir nechta overdue'larni o'chirish (faqat admin)"""
    if request.method == 'POST':
        followup_ids = request.POST.getlist('followup_ids')
        
        if not followup_ids:
            messages.error(request, "Hech qanday follow-up tanlanmagan")
        else:
            count = 0
            for followup_id in followup_ids:
                try:
                    followup = get_object_or_404(FollowUp, pk=followup_id)
                    # Mark as completed instead of deleting from database
                    followup.completed = True
                    followup.is_overdue = False
                    followup.save()
                    count += 1
                except Exception as e:
                    pass
            
            if count > 0:
                messages.success(request, f'{count} ta overdue follow-up o\'chirildi (bajarilgan deb belgilandi)')
    
    return redirect('overdue_followups_list')


# ============ TRIAL MANAGEMENT ============

@login_required
@role_required('admin', 'sales_manager', 'sales')
def trial_register(request, lead_pk):
    lead = get_object_or_404(Lead, pk=lead_pk)
    
    if request.user.is_sales and lead.assigned_sales != request.user:
        messages.error(request, "Sizda bu lidni sinovga yozish huquqi yo'q")
        return redirect('leads_list')
    
    if request.method == 'POST':
        form = TrialLessonForm(request.POST)
        if form.is_valid():
            trial = form.save(commit=False)
            trial.lead = lead
            
            # Overbooking tekshirish
            if not GroupService.can_enroll_to_group(trial.group):
                messages.error(request, 'Bu guruh to\'lgan!')
                return render(request, 'trials/register.html', {
                    'lead': lead,
                    'form': form
                })
            
            trial.save()
            
            # Guruhning trial_students ga qo'shish
            trial.group.trial_students.add(lead)
            
            lead.status = 'trial_registered'
            
            # Agar lid uchun kurs tanlanmagan bo'lsa, guruh kursini avtomatik tanlash
            if not lead.interested_course and trial.group.course:
                lead.interested_course = trial.group.course
            
            lead.save()
            messages.success(request, 'Sinovga yozildi')
            return redirect('lead_detail', pk=lead_pk)
    else:
        form = TrialLessonForm()
        # Barcha mavjud guruhlarni ko'rsatish
        # Agar interested_course bo'lsa, avval o'sha kurs guruhlarini ko'rsatish
        if lead.interested_course:
            available_groups = GroupService.get_available_groups(course=lead.interested_course)
            # Agar o'sha kursda guruhlar bo'lmasa, barcha mavjud guruhlarni ko'rsatish
            if available_groups.exists():
                form.fields['group'].queryset = available_groups
            else:
                # Barcha mavjud guruhlarni ko'rsatish
                form.fields['group'].queryset = GroupService.get_available_groups()
        else:
            # Barcha mavjud guruhlarni ko'rsatish
            form.fields['group'].queryset = GroupService.get_available_groups()
        
        # Agar hech qanday guruh bo'lmasa, barcha faol guruhlarni ko'rsatish (to'liq bo'lsa ham)
        if not form.fields['group'].queryset.exists():
            form.fields['group'].queryset = Group.objects.filter(is_active=True).order_by('name')
    
    return render(request, 'trials/register.html', {
        'lead': lead,
        'form': form
    })


@login_required
@role_required('admin', 'sales_manager', 'sales')
def trial_result(request, trial_pk):
    trial = get_object_or_404(TrialLesson, pk=trial_pk)
    
    if request.user.is_sales and trial.lead.assigned_sales != request.user:
        messages.error(request, "Sizda bu sinov natijasini kiritish huquqi yo'q")
        return redirect('leads_list')
    
    if request.method == 'POST':
        form = TrialResultForm(request.POST, instance=trial)
        if form.is_valid():
            trial = form.save()
            lead = trial.lead
            
            if trial.result == 'attended':
                lead.status = 'trial_attended'
            elif trial.result == 'not_attended':
                lead.status = 'trial_not_attended'
            elif trial.result == 'accepted':
                lead.status = 'enrolled'
                lead.enrolled_group = trial.group
                
                # Guruhning current_students oshirish
                trial.group.current_students += 1
                trial.group.save()
                
                # Trial students dan olib tashlash (agar mavjud bo'lsa)
                trial.group.trial_students.remove(lead)
            elif trial.result == 'rejected':
                lead.status = 'lost'
            
            lead.save()
            messages.success(request, 'Sinov natijasi saqlandi')
            return redirect('lead_detail', pk=lead.pk)
    else:
        form = TrialResultForm(instance=trial)
    
    return render(request, 'trials/result.html', {
        'trial': trial,
        'form': form
    })


# ============ COURSE MANAGEMENT (Admin only) ============

@login_required
@admin_required
def courses_list(request):
    courses = Course.objects.select_related().all().order_by('-created_at')
    return render(request, 'courses/list.html', {'courses': courses})


@login_required
@admin_required
def course_create(request):
    try:
        if request.method == 'POST':
            form = CourseForm(request.POST)
            if form.is_valid():
                form.save()
                messages.success(request, 'Kurs qo\'shildi')
                return redirect('courses_list')
        else:
            form = CourseForm()
        
        return render(request, 'courses/create.html', {'form': form})
    except Exception as e:
        messages.error(request, f'Xatolik yuz berdi: {str(e)}')
        import traceback
        traceback.print_exc()
        return redirect('courses_list')


@login_required
@admin_required
def course_edit(request, pk):
    try:
        course = get_object_or_404(Course, pk=pk)
        
        if request.method == 'POST':
            form = CourseForm(request.POST, instance=course)
            if form.is_valid():
                form.save()
                messages.success(request, 'Kurs yangilandi')
                return redirect('courses_list')
        else:
            form = CourseForm(instance=course)
        
        return render(request, 'courses/edit.html', {'form': form, 'course': course})
    except Exception as e:
        messages.error(request, f'Xatolik yuz berdi: {str(e)}')
        import traceback
        traceback.print_exc()
        return redirect('courses_list')


@login_required
@admin_required
def course_delete(request, pk):
    try:
        course = get_object_or_404(Course, pk=pk)
        
        if request.method == 'POST':
            course.delete()
            messages.success(request, 'Kurs o\'chirildi')
            return redirect('courses_list')
        
        return render(request, 'courses/delete.html', {'course': course})
    except Exception as e:
        messages.error(request, f'Xatolik yuz berdi: {str(e)}')
        import traceback
        traceback.print_exc()
        return redirect('courses_list')


# ============ ROOM MANAGEMENT (Admin only) ============

@login_required
@admin_required
def rooms_list(request):
    try:
        rooms = Room.objects.select_related().all().order_by('name')
        return render(request, 'rooms/list.html', {'rooms': rooms})
    except Exception as e:
        messages.error(request, f'Xatolik yuz berdi: {str(e)}')
        import traceback
        traceback.print_exc()
        return render(request, 'rooms/list.html', {'rooms': []})


@login_required
@admin_required
def room_create(request):
    if request.method == 'POST':
        form = RoomForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Xona qo\'shildi')
            return redirect('rooms_list')
    else:
        form = RoomForm()
    
    return render(request, 'rooms/create.html', {'form': form})


@login_required
@admin_required
def room_edit(request, pk):
    room = get_object_or_404(Room, pk=pk)
    
    if request.method == 'POST':
        form = RoomForm(request.POST, instance=room)
        if form.is_valid():
            form.save()
            messages.success(request, 'Xona yangilandi')
            return redirect('rooms_list')
    else:
        form = RoomForm(instance=room)
    
    return render(request, 'rooms/edit.html', {'form': form, 'room': room})


@login_required
@admin_required
def room_delete(request, pk):
    room = get_object_or_404(Room, pk=pk)
    
    if request.method == 'POST':
        room.delete()
        messages.success(request, 'Xona o\'chirildi')
        return redirect('rooms_list')
    
    return render(request, 'rooms/delete.html', {'room': room})


# ============ GROUP MANAGEMENT (Admin only) ============

@login_required
@admin_required
def groups_list(request):
    groups = Group.objects.all().select_related('course', 'room').order_by('-created_at')
    
    # Pagination qo'shish
    paginator = Paginator(groups, 25)  # 25 ta per page
    page = request.GET.get('page', 1)
    
    try:
        groups_page = paginator.page(page)
    except PageNotAnInteger:
        groups_page = paginator.page(1)
    except EmptyPage:
        groups_page = paginator.page(paginator.num_pages)
    
    return render(request, 'groups/list.html', {
        'groups': groups_page,  # Paginated
        'paginator': paginator,
        'page_obj': groups_page,
    })


@login_required
@admin_required
def group_create(request):
    try:
        if request.method == 'POST':
            form = GroupForm(request.POST)
            if form.is_valid():
                form.save()
                messages.success(request, 'Guruh qo\'shildi')
                return redirect('groups_list')
        else:
            form = GroupForm()
        
        return render(request, 'groups/create.html', {'form': form})
    except Exception as e:
        messages.error(request, f'Xatolik yuz berdi: {str(e)}')
        import traceback
        traceback.print_exc()
        return redirect('groups_list')


@login_required
@admin_required
def group_edit(request, pk):
    try:
        group = get_object_or_404(Group, pk=pk)
        
        if request.method == 'POST':
            form = GroupForm(request.POST, instance=group)
            if form.is_valid():
                form.save()
                messages.success(request, 'Guruh yangilandi')
                return redirect('groups_list')
        else:
            form = GroupForm(instance=group)
        
        return render(request, 'groups/edit.html', {'form': form, 'group': group})
    except Exception as e:
        messages.error(request, f'Xatolik yuz berdi: {str(e)}')
        import traceback
        traceback.print_exc()
        return redirect('groups_list')


@login_required
@admin_required
def group_delete(request, pk):
    try:
        group = get_object_or_404(Group, pk=pk)
        
        if request.method == 'POST':
            group.delete()
            messages.success(request, 'Guruh o\'chirildi')
            return redirect('groups_list')
        
        return render(request, 'groups/delete.html', {'group': group})
    except Exception as e:
        messages.error(request, f'Xatolik yuz berdi: {str(e)}')
        import traceback
        traceback.print_exc()
        return redirect('groups_list')


# ============ USER MANAGEMENT ============

# Sales Manager va Admin uchun
@login_required
@manager_or_admin_required
def sales_list(request):
    try:
        sales_users = User.objects.filter(role='sales').select_related().order_by('-created_at')
        
        # Pagination qo'shish
        paginator = Paginator(sales_users, 20)  # 20 ta per page
        page = request.GET.get('page', 1)
        
        try:
            sales_page = paginator.page(page)
        except PageNotAnInteger:
            sales_page = paginator.page(1)
        except EmptyPage:
            sales_page = paginator.page(paginator.num_pages)
        
        return render(request, 'users/sales_list.html', {
            'sales_users': sales_page,  # Paginated
            'paginator': paginator,
            'page_obj': sales_page,
        })
    except Exception as e:
        messages.error(request, f'Xatolik yuz berdi: {str(e)}')
        import traceback
        traceback.print_exc()
        return render(request, 'users/sales_list.html', {'sales_users': []})


# Admin uchun Sales Managerlar ro'yxati
@login_required
@admin_required
def managers_list(request):
    try:
        managers = User.objects.filter(role='sales_manager').select_related().order_by('-created_at')
        return render(request, 'users/managers_list.html', {'managers': managers})
    except Exception as e:
        messages.error(request, f'Xatolik yuz berdi: {str(e)}')
        import traceback
        traceback.print_exc()
        return render(request, 'users/managers_list.html', {'managers': []})


# Sales Manager yoki Admin uchun - Sotuvchi qo'shish
@login_required
@manager_or_admin_required
def sales_create(request):
    if request.method == 'POST':
        form = UserCreateForm(request.POST)
        if form.is_valid():
            user = form.save(commit=False)
            user.role = 'sales'  # Faqat sales rolini qo'shish mumkin
            user.save()
            messages.success(request, f'Sotuvchi {user.username} qo\'shildi')
            return redirect('sales_list')
    else:
        form = UserCreateForm(initial={'role': 'sales'})
        form.fields['role'].widget.attrs['disabled'] = True  # Role o'zgartirib bo'lmaydi
    
    return render(request, 'users/sales_create.html', {'form': form})


# Admin uchun - Sales Manager qo'shish
@login_required
@admin_required
def manager_create(request):
    if request.method == 'POST':
        form = UserCreateForm(request.POST)
        if form.is_valid():
            user = form.save(commit=False)
            user.role = 'sales_manager'  # Faqat sales_manager rolini qo'shish mumkin
            user.save()
            messages.success(request, f'Sales Manager {user.username} qo\'shildi')
            return redirect('managers_list')
    else:
        form = UserCreateForm(initial={'role': 'sales_manager'})
        form.fields['role'].widget.attrs['disabled'] = True
    
    return render(request, 'users/manager_create.html', {'form': form})


# Sales Manager yoki Admin uchun - Sotuvchi tahrirlash
@login_required
@manager_or_admin_required
def sales_edit(request, pk):
    sales = get_object_or_404(User, pk=pk, role='sales')
    
    if request.method == 'POST':
        form = UserEditForm(request.POST, instance=sales)
        if form.is_valid():
            sales_obj = form.save(commit=False)
            sales_obj.role = 'sales'  # rol o'zgarmasin
            sales_obj.save()
            form.save_m2m()  # ManyToMany ni saqlash (assigned_courses uchun)
            messages.success(request, 'Sotuvchi yangilandi')
            return redirect('sales_list')
    else:
        form = UserEditForm(instance=sales)
        # Sales Manager faqat o'z rolini o'zgartira olmaydi
        if request.user.is_sales_manager:
            form.fields['role'].widget.attrs['disabled'] = True
    
    # Follow-up statistikalarini qo'shish
    from datetime import datetime, timedelta
    from django.db.models import F
    
    today = timezone.now().date()
    now = timezone.now()
    
    # Bugungi follow-up statistikasi
    today_followups = FollowUp.objects.filter(
        sales=sales,
        due_date__date=today
    )
    today_stats = {
        'total': today_followups.count(),
        'completed': today_followups.filter(completed=True).count(),
        'pending': today_followups.filter(completed=False, is_overdue=False).count(),
        'overdue': today_followups.filter(completed=False, is_overdue=True).count(),
        'completion_rate': (today_followups.filter(completed=True).count() / today_followups.count() * 100) if today_followups.count() > 0 else 0,
    }
    
    # Oylik follow-up statistikasi
    current_month_start = datetime(today.year, today.month, 1).date()
    monthly_followups = FollowUp.objects.filter(
        sales=sales,
        due_date__date__gte=current_month_start,
        due_date__date__lte=today
    )
    monthly_stats = {
        'total': monthly_followups.count(),
        'completed': monthly_followups.filter(completed=True).count(),
        'pending': monthly_followups.filter(completed=False, is_overdue=False).count(),
        'overdue': monthly_followups.filter(completed=False, is_overdue=True).count(),
        'completion_rate': (monthly_followups.filter(completed=True).count() / monthly_followups.count() * 100) if monthly_followups.count() > 0 else 0,
        'on_time': FollowUp.objects.filter(
            sales=sales,
            completed=True,
            completed_at__lte=F('due_date'),
            due_date__date__gte=current_month_start,
            due_date__date__lte=today
        ).count(),
        'late': FollowUp.objects.filter(
            sales=sales,
            completed=True,
            completed_at__gt=F('due_date'),
            due_date__date__gte=current_month_start,
            due_date__date__lte=today
        ).count(),
    }
    
    # Overdue statistikasi (prioritet bo'yicha)
    overdue_summary = FollowUpService.get_sales_overdue_summary(sales)
    
    # Oxirgi 7 kunlik follow-up statistikasi
    last_7_days_stats = []
    for i in range(7):
        date = today - timedelta(days=i)
        day_followups = FollowUp.objects.filter(
            sales=sales,
            due_date__date=date
        )
        if day_followups.exists():
            last_7_days_stats.append({
                'date': date,
                'total': day_followups.count(),
                'completed': day_followups.filter(completed=True).count(),
                'overdue': day_followups.filter(completed=False, is_overdue=True).count(),
                'completion_rate': (day_followups.filter(completed=True).count() / day_followups.count() * 100) if day_followups.count() > 0 else 0,
            })
    
    context = {
        'form': form,
        'sales': sales,
        'today_stats': today_stats,
        'monthly_stats': monthly_stats,
        'overdue_summary': overdue_summary,
        'last_7_days_stats': last_7_days_stats,
    }
    
    return render(request, 'users/sales_edit.html', context)


# Admin uchun - Sales Manager tahrirlash
@login_required
@admin_required
def manager_edit(request, pk):
    manager = get_object_or_404(User, pk=pk, role='sales_manager')
    
    if request.method == 'POST':
        form = UserEditForm(request.POST, instance=manager)
        if form.is_valid():
            form.save()
            messages.success(request, 'Sales Manager yangilandi')
            return redirect('managers_list')
    else:
        form = UserEditForm(instance=manager)
    
    return render(request, 'users/manager_edit.html', {'form': form, 'manager': manager})


@login_required
@role_required('sales_manager')
def manager_self_edit(request):
    """Sales manager o'z profilini tahrirlashi"""
    manager = request.user
    if request.method == 'POST':
        form = UserEditForm(request.POST, instance=manager)
        if form.is_valid():
            mgr = form.save(commit=False)
            mgr.role = 'sales_manager'  # rol o'zgarmasin
            mgr.save()
            messages.success(request, 'Profil yangilandi')
            return redirect('dashboard')
    else:
        form = UserEditForm(instance=manager)
        form.fields['role'].widget.attrs['disabled'] = True
    return render(request, 'users/manager_edit.html', {'form': form, 'manager': manager})


# Sales Manager yoki Admin uchun - Sotuvchi o'chirish
@login_required
@manager_or_admin_required
def sales_delete(request, pk):
    sales = get_object_or_404(User, pk=pk, role='sales')
    
    if request.method == 'POST':
        username = sales.username
        sales.delete()
        messages.success(request, f'Sotuvchi {username} o\'chirildi')
        return redirect('sales_list')
    
    return render(request, 'users/sales_delete.html', {'sales': sales})


# Admin uchun - Sales Manager o'chirish
@login_required
@admin_required
def manager_delete(request, pk):
    manager = get_object_or_404(User, pk=pk, role='sales_manager')
    
    if request.method == 'POST':
        username = manager.username
        manager.delete()
        messages.success(request, f'Sales Manager {username} o\'chirildi')
        return redirect('managers_list')
    
    return render(request, 'users/manager_delete.html', {'manager': manager})


# ============ OFFERS ============

@login_required
def offers_list(request):
    """Takliflar ro'yxati:
    - Sales: faqat faol va amaldagi takliflar
    - Admin/Manager: barcha takliflar, filter va CRUD tugmalari
    """
    channel = request.GET.get('channel') or 'all'
    audience = request.GET.get('audience') or 'all'
    course_id = request.GET.get('course')
    course = Course.objects.filter(id=course_id).first() if course_id else None

    if request.user.is_admin or request.user.is_sales_manager:
        offers = OfferService.active_offers(
            channel=channel,
            audience=audience,
            course=course
        ) if request.GET else Offer.objects.all().order_by('-priority', '-created_at')
    else:
        offers = OfferService.active_offers(
            channel=channel,
            audience=audience,
            course=course
        )

    # Pagination qo'shish
    paginator = Paginator(offers, 20)  # 20 ta per page
    page = request.GET.get('page', 1)
    
    try:
        offers_page = paginator.page(page)
    except PageNotAnInteger:
        offers_page = paginator.page(1)
    except EmptyPage:
        offers_page = paginator.page(paginator.num_pages)
    
    context = {
        'offers': offers_page,  # Paginated
        'paginator': paginator,
        'page_obj': offers_page,
        'is_admin_or_manager': request.user.is_admin or request.user.is_sales_manager,
        'channel': channel,
        'audience': audience,
        'course_id': course_id,
        'courses': Course.objects.all(),
    }
    return render(request, 'offers/list.html', context)


@login_required
@manager_or_admin_required
def offer_create(request):
    if request.method == 'POST':
        form = OfferForm(request.POST)
        if form.is_valid():
            offer = form.save(commit=False)
            offer.created_by = request.user
            offer.save()
            messages.success(request, 'Taklif yaratildi')
            return redirect('offers_list')
    else:
        form = OfferForm()
    return render(request, 'offers/form.html', {'form': form, 'title': 'Taklif yaratish'})


@login_required
@manager_or_admin_required
def offer_edit(request, pk):
    offer = get_object_or_404(Offer, pk=pk)
    if request.method == 'POST':
        form = OfferForm(request.POST, instance=offer)
        if form.is_valid():
            form.save()
            messages.success(request, 'Taklif yangilandi')
            return redirect('offers_list')
    else:
        form = OfferForm(instance=offer)
    return render(request, 'offers/form.html', {'form': form, 'title': 'Taklif tahrirlash', 'offer': offer})


@login_required
@manager_or_admin_required
def offer_delete(request, pk):
    offer = get_object_or_404(Offer, pk=pk)
    if request.method == 'POST':
        offer.delete()
        messages.success(request, 'Taklif o\'chirildi')
        return redirect('offers_list')
    return render(request, 'offers/delete.html', {'offer': offer})


# ============ ANALYTICS ============

@login_required
@manager_or_admin_required
def analytics(request):
    from datetime import datetime, time as dt_time
    from collections import defaultdict
    
    context = {}
    today = timezone.now().date()
    now = timezone.now()
    
    # Lid statistikasi (cached)
    from django.core.cache import cache
    cache_key = f'leads_stats_{today}'
    leads_stats = cache.get(cache_key)
    
    if leads_stats is None:
        leads_stats = {
        'total': Lead.objects.count(),
        'new_today': Lead.objects.filter(created_at__date=today).count(),
        'by_status': dict(Lead.objects.values('status').annotate(count=Count('id')).values_list('status', 'count')),
        'by_source': dict(Lead.objects.values('source').annotate(count=Count('id')).values_list('source', 'count')),
    }
        # Cache for 1 hour
        cache.set(cache_key, leads_stats, 3600)
    
    context['leads_stats'] = leads_stats
    
    # Sotuvchi statistikasi (Admin va Manager uchun) - aniq statistikalar
    sales_stats = []
    for sales in User.objects.filter(role='sales', is_active_sales=True):
        kpi = KPIService.calculate_daily_kpi(sales, today)
        
        # Sotuvchi tomonidan jalb qilingan mijozlar soni
        leads_assigned = Lead.objects.filter(assigned_sales=sales).count()
        
        # Sotuvchi tomonidan sotilgan (enrolled) lidlar soni
        sales_count = Lead.objects.filter(
            assigned_sales=sales,
            status='enrolled'
        ).count()
        
        # Sotuvchi tomonidan sinovga yozilgan lidlar soni
        trials_registered = TrialLesson.objects.filter(
            lead__assigned_sales=sales
        ).select_related('lead', 'group').count()
        
        # Oylik KPI (o'rtacha)
        from datetime import datetime
        current_month_start = datetime(today.year, today.month, 1).date()
        monthly_kpis = KPI.objects.filter(
            sales=sales,
            date__gte=current_month_start,
            date__lte=today
        )
        
        # Reyting (conversion_rate bo'yicha)
        ranking = KPIService.get_sales_ranking(sales, period='month', metric='conversion_rate')
        
        sales_stats.append({
            'sales': sales,
            'kpi': kpi,
            'overdue': FollowUpService.get_overdue_followups(sales).count(),
            'leads_assigned': leads_assigned,  # Jalb qilingan mijozlar
            'sales_count': sales_count,  # Sotuvlar
            'trials_registered': trials_registered,  # Sinovga yozilganlar
            'monthly_avg_contacts': monthly_kpis.aggregate(avg=Avg('daily_contacts'))['avg'] or 0 if monthly_kpis.exists() else 0,
            'monthly_avg_conversion': monthly_kpis.aggregate(avg=Avg('conversion_rate'))['avg'] or 0 if monthly_kpis.exists() else 0,
            'monthly_total_sales': monthly_kpis.aggregate(sum=Sum('trials_to_sales'))['sum'] or 0 if monthly_kpis.exists() else 0,
            'ranking': ranking,
        })
    
    # Reyting bo'yicha tartiblash
    sales_stats.sort(key=lambda x: x['ranking']['value'] if x['ranking'] else 0, reverse=True)
    
    # Reyting raqamlarini qo'shish
    for idx, stat in enumerate(sales_stats):
        stat['rank'] = idx + 1
    
    context['sales_stats'] = sales_stats
    
    # Guruh statistikasi
    all_groups = Group.objects.filter(is_active=True)
    context['groups_stats'] = {
        'total': all_groups.count(),
        'full': all_groups.filter(current_students__gte=F('capacity')).count(),
        'available': all_groups.filter(current_students__lt=F('capacity')).count(),
    }
    
    # Xonalar statistikasi va bo'sh vaqtlar
    # O'quv markaz ish vaqti: 8:30 - 19:00 (10.5 soat = 630 minut)
    # Har bir guruh darsi: 90 minut
    # Bir kunda bir xonada maksimal guruhlar soni: 630 / 90 = 7 guruh
    WORK_START = dt_time(8, 30)  # 8:30
    WORK_END = dt_time(19, 0)    # 19:00
    LESSON_DURATION_MINUTES = 90  # 90 minut
    WORK_DURATION_MINUTES = (19 * 60) - (8 * 60 + 30)  # 630 minut
    MAX_GROUPS_PER_ROOM_PER_DAY = WORK_DURATION_MINUTES // LESSON_DURATION_MINUTES  # 7 guruh
    
    rooms = Room.objects.filter(is_active=True)
    rooms_stats = []
    total_center_capacity = 0
    total_center_available = 0
    total_center_occupied = 0
    
    # Joriy vaqt uchun bo'sh xonalar va vaqtlar
    current_time = now.time()
    current_weekday = now.weekday()  # 0 = Monday, 6 = Sunday
    
    # Hafta kunlarini Group modelidagi days formatiga moslashtirish
    day_mapping = {
        0: 'odd',  # Monday - toq kun
        1: 'even',  # Tuesday - juft kun
        2: 'odd',  # Wednesday - toq kun
        3: 'even',  # Thursday - juft kun
        4: 'odd',  # Friday - toq kun
        5: 'even',  # Saturday - juft kun
        6: 'odd',  # Sunday - toq kun
    }
    current_day_type = day_mapping.get(current_weekday, 'odd')
    
    for room in rooms:
        # Xona sig'imi
        room_capacity = room.capacity
        
        # Xonadagi guruhlar
        groups_in_room = all_groups.filter(room=room)
        
        # Xonadagi jami o'quvchilar (faqat faol guruhlardan)
        # current_students + trial_students
        total_students = sum(g.current_students for g in groups_in_room)
        total_trial_students = sum(g.trial_students.count() for g in groups_in_room)
        total_students_with_trials = total_students + total_trial_students
        
        # Xonadagi jami sig'im (faqat faol guruhlardan)
        total_capacity_groups = sum(g.capacity for g in groups_in_room)
        
        # Xonadagi bo'sh joylar (faqat to'liq bo'lmagan guruhlardan)
        # Trial students ham hisobga olinadi
        available_spots = sum(
            max(0, g.capacity - g.current_students - g.trial_students.count()) 
            for g in groups_in_room
        )
        
        # Xona uchun maksimal sig'im (ish vaqti va dars davomiyligiga qarab)
        # Bir kunda bir xonada maksimal: xona_sig'imi * maksimal_guruhlar_soni
        room_max_capacity_per_day = room_capacity * MAX_GROUPS_PER_ROOM_PER_DAY
        
        # Xonadagi joriy band o'rinlar (faqat faol guruhlardan)
        # Trial students ham hisobga olinadi
        room_occupied = total_students_with_trials
        
        # Xonadagi joriy bo'sh o'rinlar
        room_available = room_max_capacity_per_day - room_occupied
        
        # Joriy vaqtda xonada ishlayotgan guruhlar
        # 90 minutlik dars davomiyligini hisobga olish
        current_groups = []
        for group in groups_in_room.filter(days__in=[current_day_type, 'daily']):
            group_start = group.time
            # Dars davomiyligi: 90 minut
            group_end_minutes = (group_start.hour * 60 + group_start.minute + LESSON_DURATION_MINUTES) % (24 * 60)
            group_end = dt_time(group_end_minutes // 60, group_end_minutes % 60)
            
            # Joriy vaqt dars vaqtida ekanligini tekshirish
            if group_start <= current_time < group_end or (group_end < group_start and (current_time >= group_start or current_time < group_end)):
                current_groups.append(group)
        
        # Joriy vaqtda bo'sh ekanligini tekshirish
        is_free_now = len(current_groups) == 0
        
        # Xonaning band vaqtlari (guruhlar bo'lgan vaqtlar)
        busy_times = sorted(set(groups_in_room.values_list('time', flat=True)))
        
        rooms_stats.append({
            'room': room,
            'capacity': room_capacity,
            'total_students': total_students,
            'total_trial_students': total_trial_students,
            'total_students_with_trials': total_students_with_trials,
            'total_capacity': total_capacity_groups,
            'available_spots': available_spots,
            'groups_count': groups_in_room.count(),
            'is_free_now': is_free_now,
            'busy_times': busy_times,
            'max_capacity_per_day': room_max_capacity_per_day,
            'occupied_per_day': room_occupied,
            'available_per_day': room_available,
        })
        
        # Markaz sig'imi hisoblash
        total_center_capacity += room_max_capacity_per_day
        total_center_occupied += room_occupied
        total_center_available += room_available
    
    context['rooms_stats'] = rooms_stats
    context['center_stats'] = {
        'total_capacity': total_center_capacity,
        'total_available': total_center_available,
        'total_occupied': total_center_occupied,
        'occupancy_percentage': (total_center_occupied / total_center_capacity * 100) if total_center_capacity > 0 else 0,
        'work_hours': f"{WORK_START.strftime('%H:%M')} - {WORK_END.strftime('%H:%M')}",
        'lesson_duration': LESSON_DURATION_MINUTES,
        'max_groups_per_room': MAX_GROUPS_PER_ROOM_PER_DAY,
    }
    
    # Guruhlar va vaqtlar bo'yicha statistikalar
    groups_by_time = defaultdict(list)
    for group in all_groups:
        groups_by_time[group.time].append(group)
    
    context['groups_by_time'] = dict(sorted(groups_by_time.items()))
    
    return render(request, 'analytics/index.html', context)


@login_required
@role_required('sales')
def sales_kpi(request):
    """Sotuvchi uchun o'z KPI'larini ko'rish"""
    sales = request.user
    today = timezone.now().date()
    
    # Bugungi KPI
    today_kpi = KPIService.calculate_daily_kpi(sales, today)
    
    # Oxirgi 7 kunlik KPI
    from datetime import timedelta
    last_7_days = []
    for i in range(7):
        date = today - timedelta(days=i)
        kpi = KPI.objects.filter(sales=sales, date=date).first()
        if kpi:
            last_7_days.append({
                'date': date,
                'kpi': kpi
            })
        else:
            # Agar KPI hisoblanmagan bo'lsa, hisoblaymiz
            kpi = KPIService.calculate_daily_kpi(sales, date)
            last_7_days.append({
                'date': date,
                'kpi': kpi
            })
    
    # Oylik statistikalar
    from datetime import datetime
    current_month_start = datetime(today.year, today.month, 1).date()
    monthly_kpis = KPI.objects.filter(
        sales=sales,
        date__gte=current_month_start,
        date__lte=today
    ).order_by('-date')
    
    # Oylik konversiya (berilgan lidlar -> enrolled)
    monthly_conv = KPIService.calculate_monthly_conversion_rate(sales, today.year, today.month)
    
    # Oylik yig'indilar
    monthly_stats = {
        'total_contacts': sum(kpi.daily_contacts for kpi in monthly_kpis),
        'total_followups': sum(kpi.daily_followups for kpi in monthly_kpis),
        'total_trials': sum(kpi.trials_registered for kpi in monthly_kpis),
        'total_sales': sum(kpi.trials_to_sales for kpi in monthly_kpis),
        'avg_conversion': monthly_conv['conversion_rate'],
        'avg_response_time': sum(kpi.response_time_minutes for kpi in monthly_kpis) / len(monthly_kpis) if monthly_kpis else 0,
        'monthly_assigned': monthly_conv['total_assigned'],
        'monthly_enrolled': monthly_conv['enrolled'],
    }
    
    # Mening lidlarim statistikasi
    my_leads = Lead.objects.filter(assigned_sales=sales)
    leads_stats = {
        'total': my_leads.count(),
        'new': my_leads.filter(status='new').count(),
        'contacted': my_leads.filter(status='contacted').count(),
        'interested': my_leads.filter(status='interested').count(),
        'trial_registered': my_leads.filter(status='trial_registered').count(),
        'enrolled': my_leads.filter(status='enrolled').count(),
        'lost': my_leads.filter(status='lost').count(),
    }
    
    # Oxirgi 30 kunlik KPI (batafsil)
    from datetime import timedelta
    last_30_days_kpi = []
    for i in range(30):
        date = today - timedelta(days=i)
        kpi = KPI.objects.filter(sales=sales, date=date).first()
        if not kpi:
            kpi = KPIService.calculate_daily_kpi(sales, date)
        last_30_days_kpi.append({
            'date': date,
            'kpi': kpi
        })
    
    # O'rtacha, minimum, maksimum
    if last_30_days_kpi:
        contacts_values = [day['kpi'].daily_contacts for day in last_30_days_kpi]
        conversion_values = [day['kpi'].conversion_rate for day in last_30_days_kpi]
        stats_summary = {
            'contacts': {
                'avg': sum(contacts_values) / len(contacts_values) if contacts_values else 0,
                'min': min(contacts_values) if contacts_values else 0,
                'max': max(contacts_values) if contacts_values else 0,
            },
            'conversion': {
                'avg': sum(conversion_values) / len(conversion_values) if conversion_values else 0,
                'min': min(conversion_values) if conversion_values else 0,
                'max': max(conversion_values) if conversion_values else 0,
            }
        }
    else:
        stats_summary = {
            'contacts': {'avg': 0, 'min': 0, 'max': 0},
            'conversion': {'avg': 0, 'min': 0, 'max': 0}
        }
    
    # Reyting va taqqoslash
    ranking_conversion = KPIService.get_sales_ranking(sales, period='month', metric='conversion_rate')
    ranking_contacts = KPIService.get_sales_ranking(sales, period='month', metric='daily_contacts')
    comparison_data = KPIService.get_comparison_data(sales, period='month')
    
    # Haftalik xulosa
    from datetime import datetime
    week_start = today - timedelta(days=today.weekday())
    weekly_summary = KPIService.get_weekly_kpi_summary(sales, week_start)
    
    # Follow-up statistikalarini qo'shish
    from django.db.models import F
    
    # Oylik follow-up statistikasi
    monthly_followups = FollowUp.objects.filter(
        sales=sales,
        due_date__date__gte=current_month_start,
        due_date__date__lte=today
    )
    followup_stats = {
        'total': monthly_followups.count(),
        'completed': monthly_followups.filter(completed=True).count(),
        'overdue': monthly_followups.filter(completed=False, is_overdue=True).count(),
        'on_time': FollowUp.objects.filter(
            sales=sales,
            completed=True,
            completed_at__lte=F('due_date'),
            due_date__date__gte=current_month_start,
            due_date__date__lte=today
        ).count(),
        'late': FollowUp.objects.filter(
            sales=sales,
            completed=True,
            completed_at__gt=F('due_date'),
            due_date__date__gte=current_month_start,
            due_date__date__lte=today
        ).count(),
        'completion_rate': (monthly_followups.filter(completed=True).count() / monthly_followups.count() * 100) if monthly_followups.count() > 0 else 0,
    }
    
    # Overdue statistikasi (prioritet bo'yicha)
    overdue_summary = FollowUpService.get_sales_overdue_summary(sales)
    
    context = {
        'today_kpi': today_kpi,
        'last_7_days': last_7_days,
        'last_30_days_kpi': last_30_days_kpi,
        'stats_summary': stats_summary,
        'monthly_stats': monthly_stats,
        'leads_stats': leads_stats,
        'monthly_kpis': monthly_kpis[:30],  # Oxirgi 30 kunlik
        'ranking_conversion': ranking_conversion,
        'ranking_contacts': ranking_contacts,
        'comparison_data': comparison_data,
        'weekly_summary': weekly_summary,
        'followup_stats': followup_stats,
        'overdue_summary': overdue_summary,
    }
    
    return render(request, 'analytics/sales_kpi.html', context)


@login_required
@manager_or_admin_required
def export_analytics_excel(request):
    """Analytics ma'lumotlarini Excel'ga export qilish"""
    from datetime import datetime
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from django.http import HttpResponse
    
    today = timezone.now().date()
    current_month_start = datetime(today.year, today.month, 1).date()
    
    # Workbook yaratish
    wb = Workbook()
    ws = wb.active
    ws.title = "Sotuvchilar Reytingi"
    
    # Header style
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Header qatorlar
    headers = ['Reyting', 'Sotuvchi', 'Jalb Qilingan', 'Aloqa (Kunlik)', 'Aloqa (O\'rtacha)', 
               'Qayta Aloqa', 'Sinov', 'Sotuv (Kunlik)', 'Sotuv (Oylik)', 
               'Konversiya (Kunlik)', 'Konversiya (O\'rtacha)', 'Muddati O\'tgan']
    ws.append(headers)
    
    # Header formatlash
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Ma'lumotlar
    sales_stats = []
    for sales in User.objects.filter(role='sales', is_active_sales=True):
        kpi = KPIService.calculate_daily_kpi(sales, today)
        leads_assigned = Lead.objects.filter(assigned_sales=sales).count()
        sales_count = Lead.objects.filter(assigned_sales=sales, status='enrolled').count()
        trials_registered = TrialLesson.objects.filter(lead__assigned_sales=sales).count()
        overdue = FollowUpService.get_overdue_followups(sales).count()
        
        monthly_kpis = KPI.objects.filter(
            sales=sales,
            date__gte=current_month_start,
            date__lte=today
        )
        monthly_avg_contacts = monthly_kpis.aggregate(avg=Avg('daily_contacts'))['avg'] or 0 if monthly_kpis.exists() else 0
        monthly_avg_conversion = monthly_kpis.aggregate(avg=Avg('conversion_rate'))['avg'] or 0 if monthly_kpis.exists() else 0
        monthly_total_sales = monthly_kpis.aggregate(sum=Sum('trials_to_sales'))['sum'] or 0 if monthly_kpis.exists() else 0
        
        ranking = KPIService.get_sales_ranking(sales, period='month', metric='conversion_rate')
        
        sales_stats.append({
            'rank': ranking['rank'],
            'sales': sales,
            'kpi': kpi,
            'leads_assigned': leads_assigned,
            'sales_count': sales_count,
            'trials_registered': trials_registered,
            'overdue': overdue,
            'monthly_avg_contacts': monthly_avg_contacts,
            'monthly_avg_conversion': monthly_avg_conversion,
            'monthly_total_sales': monthly_total_sales,
        })
    
    # Reyting bo'yicha tartiblash
    sales_stats.sort(key=lambda x: x['rank'])
    
    # Ma'lumotlarni qo'shish
    for stat in sales_stats:
        row = [
            stat['rank'],
            stat['sales'].username,
            stat['leads_assigned'],
            stat['kpi'].daily_contacts,
            round(stat['monthly_avg_contacts'], 1),
            stat['kpi'].daily_followups,
            stat['trials_registered'],
            stat['sales_count'],
            stat['monthly_total_sales'],
            round(stat['kpi'].conversion_rate, 1),
            round(stat['monthly_avg_conversion'], 1),
            stat['overdue'],
        ]
        ws.append(row)
    
    # Column width'ni sozlash
    column_widths = [10, 20, 15, 15, 15, 15, 10, 15, 15, 15, 15, 15]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
    
    # Response yaratish
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"analytics_export_{today.strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response


# ============ LEAVE REQUESTS ============

@login_required
@role_required('sales')
def leave_request_create(request):
    """Sotuvchi tomonidan ruxsat so'rash"""
    if request.method == 'POST':
        form = LeaveRequestForm(request.POST)
        if form.is_valid():
            leave_request = form.save(commit=False)
            leave_request.sales = request.user
            leave_request.save()
            messages.success(request, 'Ruxsat so\'rovi yuborildi. Manager tasdiqlashini kutib turing.')
            return redirect('leave_request_list')
    else:
        form = LeaveRequestForm()
    
    return render(request, 'leaves/create.html', {'form': form})


@login_required
@role_required('sales')
def leave_request_list(request):
    """Sotuvchi tomonidan o'z ruxsat so'rovlari ro'yxati"""
    leave_requests = LeaveRequest.objects.filter(sales=request.user).order_by('-created_at')
    
    # Kelgan xabarlar
    received_messages = SalesMessage.objects.filter(recipients=request.user).order_by('-created_at')[:5]
    read_message_ids = list(SalesMessageRead.objects.filter(
        user=request.user
    ).values_list('message_id', flat=True))
    unread_count = SalesMessage.objects.filter(
        recipients=request.user
    ).exclude(id__in=read_message_ids).count()
    
    return render(request, 'leaves/list.html', {
        'leave_requests': leave_requests,
        'received_messages': received_messages,
        'unread_count': unread_count,
        'read_message_ids': read_message_ids,
    })


@login_required
@manager_or_admin_required
def leave_request_pending_list(request):
    """Manager/Admin uchun kutilayotgan ruxsat so'rovlari"""
    pending_requests = LeaveRequest.objects.filter(status='pending').order_by('-created_at')
    return render(request, 'leaves/pending_list.html', {'pending_requests': pending_requests})


@login_required
@manager_or_admin_required
def leave_request_approve(request, pk):
    """Ruxsat so'rovini tasdiqlash yoki rad etish"""
    leave_request = get_object_or_404(LeaveRequest, pk=pk)
    
    if request.method == 'POST':
        form = LeaveRequestApprovalForm(request.POST, instance=leave_request)
        if form.is_valid():
            status = form.cleaned_data['status']
            if status == 'approved':
                leave_request.approve(request.user)
                messages.success(request, f'{leave_request.sales.username} uchun ruxsat tasdiqlandi')
            elif status == 'rejected':
                leave_request.reject(request.user, form.cleaned_data.get('rejection_reason', ''))
                messages.success(request, f'{leave_request.sales.username} uchun ruxsat rad etildi')
            return redirect('leave_request_pending_list')
    else:
        form = LeaveRequestApprovalForm(instance=leave_request)
    
    return render(request, 'leaves/approve.html', {'form': form, 'leave_request': leave_request})


# ============ SALES ABSENCE ============

@login_required
@manager_or_admin_required
def sales_absence_set(request, pk):
    """Manager tomonidan sotuvchini ishda emasligini belgilash"""
    sales = get_object_or_404(User, pk=pk, role='sales')
    
    if request.method == 'POST':
        form = SalesAbsenceForm(request.POST, instance=sales)
        if form.is_valid():
            form.save()
            if sales.is_absent:
                messages.success(request, f'{sales.username} ishda emasligi belgilandi')
            else:
                messages.success(request, f'{sales.username} ishda emasligi bekor qilindi')
            return redirect('sales_list')
    else:
        form = SalesAbsenceForm(instance=sales)
    
    return render(request, 'users/sales_absence.html', {'form': form, 'sales': sales})


# ============ SALES MESSAGES ============

@login_required
@manager_or_admin_required
def sales_message_create(request):
    """Sotuvchilarga xabar yuborish"""
    if request.method == 'POST':
        form = SalesMessageForm(request.POST)
        if form.is_valid():
            message = form.save(commit=False)
            message.sender = request.user
            message.save()
            form.save_m2m()  # ManyToMany field'larni saqlash
            
            # Telegram orqali xabar yuborish
            from .telegram_bot import send_telegram_notification
            
            recipients = form.cleaned_data['recipients']
            telegram_sent_count = 0
            
            priority_emoji = {
                'urgent': '🚨',
                'high': '⚠️',
                'normal': '📢',
                'low': 'ℹ️',
            }.get(message.priority, '📢')
            
            for recipient in recipients:
                # Telegram orqali yuborish
                if recipient.telegram_chat_id:
                    telegram_message = (
                        f"{priority_emoji} {message.subject}\n\n"
                        f"{message.message}\n\n"
                        f"Yuboruvchi: {message.sender.username}"
                    )
                    
                    if send_telegram_notification(recipient.telegram_chat_id, telegram_message):
                        telegram_sent_count += 1
            
            message.telegram_sent = telegram_sent_count > 0
            message.save()
            
            messages.success(
                request, 
                f'Xabar {recipients.count()} ta sotuvchiga yuborildi. '
                f'Telegram orqali: {telegram_sent_count} ta'
            )
            return redirect('sales_message_list')
    else:
        form = SalesMessageForm()
    
    return render(request, 'messages/create.html', {'form': form})


@login_required
@manager_or_admin_required
def sales_message_list(request):
    """Yuborilgan xabarlar ro'yxati"""
    sent_messages = SalesMessage.objects.filter(sender=request.user).order_by('-created_at')
    return render(request, 'messages/list.html', {'messages': sent_messages})


@login_required
@role_required('sales')
def sales_message_inbox(request):
    """Sotuvchi uchun kelgan xabarlar"""
    received_messages = SalesMessage.objects.filter(recipients=request.user).order_by('-created_at')
    
    # O'qilgan xabarlarni tekshirish
    read_message_ids = SalesMessageRead.objects.filter(
        user=request.user
    ).values_list('message_id', flat=True)
    
    unread_count = received_messages.exclude(id__in=read_message_ids).count()
    
    return render(request, 'messages/inbox.html', {
        'inbox_messages': received_messages,
        'unread_count': unread_count,
        'read_message_ids': read_message_ids,
    })


@login_required
@role_required('sales')
def sales_message_detail(request, pk):
    """Xabar tafsilotlari va o'qilganligini belgilash"""
    message = get_object_or_404(SalesMessage, pk=pk)
    
    # Faqat xabar qabul qiluvchisi ko'ra oladi
    if request.user not in message.recipients.all():
        messages.error(request, "Sizda bu xabarni ko'rish huquqi yo'q")
        return redirect('sales_message_inbox')
    
    # Xabarni o'qilgan deb belgilash
    SalesMessageRead.objects.get_or_create(
        message=message,
        user=request.user
    )
    
    return render(request, 'messages/detail.html', {'message': message})


@login_required
def sales_message_delete(request, pk):
    """Xabarni o'chirish (admin yoki yuboruvchi)"""
    message = get_object_or_404(SalesMessage, pk=pk)
    
    # Faqat admin yoki yuboruvchi o'chira oladi
    if not (request.user.is_admin or message.sender == request.user):
        messages.error(request, "Sizda bu xabarni o'chirish huquqi yo'q")
        return redirect('sales_message_list')
    
    if request.method == 'POST':
        message_subject = message.subject
        message.delete()
        messages.success(request, f'Xabar "{message_subject}" o\'chirildi')
        
        # Agar admin bo'lsa, list sahifasiga, aks holda o'z list sahifasiga
        if request.user.is_admin or request.user.role == 'sales_manager':
            return redirect('sales_message_list')
        else:
            return redirect('sales_message_inbox')
    
    return render(request, 'messages/delete_confirm.html', {'message': message})