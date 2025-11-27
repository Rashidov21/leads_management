"""Excel and CSV file parsing service."""
from typing import List, Dict
from openpyxl import load_workbook
import csv
import io


class ExcelService:
    """Service for parsing Excel files."""
    
    @staticmethod
    def parse_file(file_obj) -> List[Dict]:
        """
        Parse Excel file and extract leads.
        
        Expects:
            Column A: Name
            Column B: Phone
        
        Args:
            file_obj: File object
        
        Returns:
            List of dictionaries with 'name' and 'phone' keys
        """
        leads = []
        
        try:
            workbook = load_workbook(file_obj)
            worksheet = workbook.active
            
            # Skip header row if present
            for idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, values_only=True)):
                if len(row) >= 2:
                    name = str(row[0]).strip() if row[0] else ''
                    phone = str(row[1]).strip() if row[1] else ''
                    
                    if name and phone:
                        leads.append({
                            'name': name,
                            'phone': phone
                        })
        except Exception as e:
            print(f"Error parsing Excel file: {e}")
            return []
        
        return leads


class CSVService:
    """Service for parsing CSV files."""
    
    @staticmethod
    def parse_file(file_obj, delimiter=',') -> List[Dict]:
        """
        Parse CSV file and extract leads.
        
        Expects columns:
            Column 1: Name
            Column 2: Phone
        
        Args:
            file_obj: File object
            delimiter: CSV delimiter character
        
        Returns:
            List of dictionaries with 'name' and 'phone' keys
        """
        leads = []
        
        try:
            # Handle both binary and text mode
            if hasattr(file_obj, 'read'):
                content = file_obj.read()
                if isinstance(content, bytes):
                    content = content.decode('utf-8')
                file_obj = io.StringIO(content)
            
            reader = csv.reader(file_obj, delimiter=delimiter)
            
            # Skip header row
            next(reader, None)
            
            for row in reader:
                if len(row) >= 2:
                    name = row[0].strip() if row[0] else ''
                    phone = row[1].strip() if row[1] else ''
                    
                    if name and phone:
                        leads.append({
                            'name': name,
                            'phone': phone
                        })
        except Exception as e:
            print(f"Error parsing CSV file: {e}")
            return []
        
        return leads
